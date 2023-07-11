from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import pytest
from constants import globalConstants
import openpyxl
from pathlib import Path
from datetime import date

class Test_SauceDemo:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(globalConstants.url)
        self.driver.maximize_window()
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        

    def teardown_method(self):
        self.driver.quit()
    
    def getData():
        excelFile = openpyxl.load_workbook("data/success_login.xlsx") 
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        
        return data


    #başarılı giriş
    @pytest.mark.parametrize("username,password",getData())
    def test_valid_login(self,username,password):

        self.waitForElementVisible((By.ID,globalConstants.userNameID))
        userNameInput = self.driver.find_element(By.ID,globalConstants.userNameID)
        userNameInput.send_keys(username)

        
        self.waitForElementVisible((By.ID,globalConstants.passwordID))
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordID)
        passwordInput.send_keys(password)
        logInBtn = self.driver.find_element(By.ID,globalConstants.loginBtnID)
        logInBtn.click()
        
        self.waitForElementVisible((By.XPATH,globalConstants.filtreleXpath))
        filtrele = self.driver.find_element(By.XPATH,globalConstants.filtreleXpath)
        filtrele.click()
        
        self.waitForElementVisible((By.XPATH,globalConstants.yuksekFiyatXpath))
        yuksekFiyat = self.driver.find_element(By.XPATH,globalConstants.yuksekFiyatXpath)
        yuksekFiyat.click()
        
        self.waitForElementVisible((By.ID,globalConstants.addToCartID))
        addToCart = self.driver.find_element(By.ID,globalConstants.addToCartID)
        addToCart.click()
        
        self.waitForElementVisible((By.ID,globalConstants.addToCartID_2))
        addToCart_2 = self.driver.find_element(By.ID,globalConstants.addToCartID_2)
        addToCart_2.click()
        
        self.waitForElementVisible((By.XPATH,globalConstants.goToBasketXpath))
        goToBasket = self.driver.find_element(By.XPATH,globalConstants.goToBasketXpath)
        goToBasket.click()
        
        self.waitForElementVisible((By.ID,globalConstants.checkoutID))
        checkout = self.driver.find_element(By.ID,globalConstants.checkoutID)
        checkout.click()
        
        self.waitForElementVisible((By.ID,globalConstants.firstNameID))
        firstName = self.driver.find_element(By.ID,globalConstants.firstNameID)
        firstName.send_keys("halit")
        lastName = self.driver.find_element(By.ID,globalConstants.lastNameID)
        lastName.send_keys("kinik")
        postalCode = self.driver.find_element(By.ID,globalConstants.postalCodeID)
        postalCode.send_keys("12345")
        
        continueBasket = self.driver.find_element(By.ID,globalConstants.continueBasket) 
        continueBasket.click()
        
        self.waitForElementVisible((By.ID,globalConstants.finishID))
        finish = self.driver.find_element(By.ID,globalConstants.finishID)
        finish.click()
        
        self.waitForElementVisible((By.ID,globalConstants.backToHome))
        backToHome = self.driver.find_element(By.ID,globalConstants.backToHome)
        backToHome.click()
        
        self.waitForElementVisible((By.ID,globalConstants.menuBtnID))
        menuBtn = self.driver.find_element(By.ID,globalConstants.menuBtnID)
        menuBtn.click()
        
        self.waitForElementVisible((By.ID,globalConstants.logoutID))
        logOut = self.driver.find_element(By.ID,globalConstants.logoutID)
        logOut.click()


    #username dogru password yanlıs olduğunda
    @pytest.mark.parametrize("username,password",[("standard_user","123456789")])
    def test_false_password(self,username,password):

        self.waitForElementVisible((By.ID,globalConstants.userNameID))
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameID)
        usernameInput.send_keys(username)

        self.waitForElementVisible((By.ID,globalConstants.passwordID))
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordID)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,globalConstants.loginBtnID)
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,globalConstants.errorMessageXpath) 

        self.driver.save_screenshot(f"{self.folderPath}/tes-false-password-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    #username yanlış password doğru olduğunda
    @pytest.mark.parametrize("username,password",[("standard_userr","secret_sauce")])
    def test_false_username(self,username,password):
        self.waitForElementVisible((By.ID,globalConstants.userNameID))
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameID)
        usernameInput.send_keys(username)

        self.waitForElementVisible((By.ID,globalConstants.passwordID))
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordID)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,globalConstants.loginBtnID)
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,globalConstants.errorMessageXpath)
        self.driver.save_screenshot(f"{self.folderPath}/tes-false-username-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"    
    
    #ikisi de yanlış olduğunda
    @pytest.mark.parametrize("username,password",[("standard_userrr","secret_sauceee")])
    def test_false_both(self,username,password):
        self.waitForElementVisible((By.ID,globalConstants.userNameID))
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameID)
        usernameInput.send_keys(username)

        self.waitForElementVisible((By.ID,globalConstants.passwordID))
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordID)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,globalConstants.loginBtnID)
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,globalConstants.errorMessageXpath)
        self.driver.save_screenshot(f"{self.folderPath}/tes-false-both-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"    


    #ikisi de boş geçildiğinde 
    @pytest.mark.parametrize("username,password",[("","")])
    def test_empty_username_password(self,username,password):
        self.waitForElementVisible((By.ID,globalConstants.userNameID))
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameID)
        usernameInput.send_keys(username)

        self.waitForElementVisible((By.ID,globalConstants.passwordID))
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordID)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,globalConstants.loginBtnID)
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,globalConstants.errorMessageXpath) 
        self.driver.save_screenshot(f"{self.folderPath}/tes-empty-username-password-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username is required"    

    def waitForElementVisible(self,locator,timeout=10):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
    